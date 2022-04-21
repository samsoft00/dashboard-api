from mysql.connector import connect, Error
from openpyxl import load_workbook
from requests.models import HTTPError
from datetime import datetime
import requests
import pendulum
import traceback
import json
import sys
import re

#validate url
regex = re.compile(
        r'^(?:http|ftp)s?://' # http:// or https://
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|' #domain...
        r'localhost|' #localhost...
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})' # ...or ip
        r'(?::\d+)?' # optional port
        r'(?:/?|[/?]\S+)$', re.IGNORECASE)


columns = {
    'name': 'C',
    'reference_no': 'D',
    'trans_type': 'E',
    'currency': 'F',
    'rate': 'V',
    'total_purchased': 'W',
    'total_sold': 'X',
    'payment_mode': 'Y',
    'cash_received': 'Z',
    'cash_paid': 'AA',
    'cash_balance': 'AB',
    'paying_bank': 'AC',
    'receiving_bank': 'AD',
    'status': 'AE',
    'usd': {
        'purchase': 'G',
        'sold': 'H',
        'balance': 'I',
    },
    'gbp': {
        'purchase': 'K',
        'sold': 'L',
        'balance': 'M',
    },
    'euro': {
        'purchase': 'O',
        'sold': 'P',
        'balance': 'Q',
    },
    'uk_gbp': {
        'purchase': 'S',
        'sold': 'T',
        'balance': 'U',
    },
}

# BdcReportGenerator
class BdcReportGenerator:

    def __init__(self, dbcred, file_path, write_path):
        today = pendulum.now()

        self.start_datetime = today.start_of('day')
        self.end_datetime = today.end_of('day')

        self.file_path = file_path
        self.write_path = write_path

        self.cnx = connect(
            host=dbcred.get('host'),
            database=dbcred.get('database'),
            user=dbcred.get('user'),
            password=dbcred.get('pass'),
            auth_plugin="mysql_native_password")

        self.cur = self.cnx.cursor(buffered=True, dictionary=True)

    def getBdcOrders(self):
        # select bdc orders
        select_bdc_query = """
            SELECT 
            o.id AS id, 
            o.refrence_no AS refrence_no, 
            o.customer->>'$.name' AS customer,
            o.transaction_type AS transaction_type, 
            o.bdc_company as bdc_company,
            c.locale AS currency,
            case 
                when o.transaction_type = 'buy' 
                then o.volume else NULL 
            end AS volume_purchased,
            case 
                when o.transaction_type = 'sell' 
                then o.volume else NULL 
            end AS volume_sold,
            o.exchange_rate AS rate,
            case 
                when o.transaction_type = 'buy' 
                then (o.volume * o.exchange_rate) else ' -   ' 
            end AS total_amount_purchased, 
            case 
                when o.transaction_type = 'sell' 
                then (o.volume * o.exchange_rate) else ' -   ' 
            end AS total_amount_sold, 
            o.mode_of_payment AS mode_of_payment, 
            case 
                when o.transaction_type = 'buy' 
                then o.cash_payment else NULL 
            end AS cash_received, 
            case 
                when o.transaction_type = 'sell' 
                then o.cash_payment else NULL 
            end AS cash_paid, 
            case 
                when mode_of_payment = 'wire' AND transaction_type = 'buy' 
                then b.account_name else '' 
            end AS receiving_bank, 
            case 
                when mode_of_payment = 'wire' AND transaction_type = 'sell' 
                then b.account_name else '' 
            end AS paying_bank,
            o.created_at AS created_at, 
            o.status AS status 
            FROM bdc_orders o 
            INNER JOIN currency_types c ON o.currency_type_id = c.id 
            LEFT JOIN bdc_bank_details b ON o.bdc_bank_detail_id = b.id 
            WHERE o.created_at BETWEEN %s AND %s 
            ORDER BY o.created_at DESC
        """

        self.cur.execute(select_bdc_query, (self.start_datetime, self.end_datetime))
        return self.cur.fetchall()
    
    def getStockBalances(self):
        # select stock balances
        select_stocks = """
            SELECT 
            c.name as name, 
            c.locale as locale, 
            s.stock_balance as stock_balance, 
            sb.opening_balance as opening_balance, 
            (case 
                when sb.closing_balance = '0.00' 
                then s.stock_balance else sb.closing_balance 
            end) as closing_balance,
            @volume := COALESCE((SELECT SUM(volume) FROM bdc_orders WHERE transaction_type = %(sell)s AND currency_type_id = c.id AND created_at BETWEEN %(start)s AND %(end)s), 0.00) as volume_sold, 
            @avg_selling_rate := COALESCE(( SELECT CAST(AVG(DISTINCT exchange_rate) AS DECIMAL(12,2)) FROM bdc_orders WHERE transaction_type = %(sell)s AND currency_type_id = c.id AND created_at BETWEEN %(start)s AND %(end)s), 0.00) AS average_selling_rate, 
            @avg_purchase_rate := COALESCE(( SELECT CAST(AVG(DISTINCT exchange_rate) AS DECIMAL(12,2)) FROM bdc_orders WHERE transaction_type = %(buy)s AND currency_type_id = c.id AND created_at BETWEEN %(start)s AND %(end)s), 0.00) AS average_purchase_rate, 
            @sprd := CAST(@avg_selling_rate - @avg_purchase_rate AS DECIMAL(12,2)) AS spread, 
            CAST(@sprd * @volume AS DECIMAL(12,2)) AS profit, 
            sb.created_at as created_at, 
            sb.updated_at as updated_at 
            FROM currency_types c 
            LEFT JOIN bdc_stocks s ON c.id = s.currency_type_id 
            LEFT JOIN bdc_stock_balances sb ON s.id = sb.bdc_stock_id 
            WHERE sb.created_at BETWEEN %(start)s AND %(end)s 
            ORDER BY s.id DESC
        """

        data = {
            'start': self.start_datetime,
            'end': self.end_datetime,
            'sell': 'sell',
            'buy': 'buy'
        }
        self.cur.execute(select_stocks, data)
        return self.cur.fetchall()

    def uploadFile(self):

        files = {'image': (self.write_path, open(self.write_path, 'rb'), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', {'Expires': '0'})}
        r = requests.post('https://kyc-api.cpfs.online/upload', files=files)
        
        if r.status_code != 201 or r.status_code != 200:
            r.raise_for_status()
        
        response = json.loads(r.text)
        assert('message' in response and re.match(regex,response.get('message'))), 'Unable to determine file upload URL'

        print (response.get('message'))

    def generateReport(self):
        workbook = load_workbook(self.file_path)
        current_sheet = workbook.active #return active page
        ws = workbook.copy_worksheet(current_sheet)

        orders = self.getBdcOrders()
        stocks = self.getStockBalances()

        if not len(orders) or not len(stocks):
            raise IOError('No recent activities on dashboard to export, try again after sometime.')

        stock_balances = 6
        for s in stocks:
            key = ws['C{0}'.format(stock_balances)].value.strip().upper()
            for j in stocks:
                if j['locale'] == key:
                    ws['E{0}'.format(stock_balances)] = j['average_selling_rate']
                    ws['F{0}'.format(stock_balances)] = j['average_purchase_rate']
            stock_balances += 1

        # opening balances
        stock_num = 17
        ws['I{0}'.format(stock_num)] = stocks[1]['opening_balance']  # USD
        ws['M{0}'.format(stock_num)] = stocks[2]['opening_balance']  # GBP
        ws['Q{0}'.format(stock_num)] = stocks[3]['opening_balance']  # EUR
        # ws['U{0}'.format(stock_num)] = stocks[4]['opening_balance'] # UK GBP
        ws['AB{0}'.format(stock_num)] = stocks[0]['opening_balance']  # NGN
        #
        row_start = 18
        for item in orders:
            ws['{0}{1}'.format(columns.get('name'), row_start)] = item['customer'] #.decode('utf-8')
            ws['{0}{1}'.format(columns.get('reference_no'), row_start)] = int(item['refrence_no'])
            ws['{0}{1}'.format(columns.get('trans_type'), row_start)] = item['transaction_type'].capitalize()
            ws['{0}{1}'.format(columns.get('currency'), row_start)] = item['currency']
            ws['{0}{1}'.format(columns.get('rate'), row_start)] = item['rate']
            ws['{0}{1}'.format(columns.get('payment_mode'), row_start)] = item['mode_of_payment']
            if item['cash_received'] is not None:
                ws['{0}{1}'.format(columns.get('cash_received'), row_start)] = float(item['cash_received'])
            if item['cash_paid'] is not None:
                ws['{0}{1}'.format(columns.get('cash_paid'), row_start)] = float(item['cash_paid'])
            ws['{0}{1}'.format(columns.get('paying_bank'), row_start)] = item['paying_bank']
            ws['{0}{1}'.format(columns.get('receiving_bank'), row_start)] = item['receiving_bank']
            ws['{0}{1}'.format(columns.get('status'), row_start)] = item['status']

            # check currency type
            cur_type = dict()
            currency = item['currency'].upper()

            if currency == 'EUR':
                cur_type = dict(columns.get('euro'))
            elif currency == 'GBP':
                cur_type = dict(columns.get('gbp'))
            elif currency == 'USD':
                cur_type = dict(columns.get('usd'))
            elif currency == 'UK_GBP':
                cur_type = dict(columns.get('uk_gbp'))

            if not cur_type:
                raise TypeError('Unable to determine order currency type')

            ws['{0}{1}'.format(cur_type.get('purchase'), row_start)] = item['volume_purchased']
            ws['{0}{1}'.format(cur_type.get('sold'), row_start)] = item['volume_sold']
            row_start += 1

        ws.title = datetime.today().strftime('%d-%m-%Y')
        workbook.remove_sheet(current_sheet)
        workbook.save(self.write_path)

        workbook.close()

        self.uploadFile()

def generate_report():
    try:
        dbcred = dict({ 
            'host': sys.argv[1], 
            'database': sys.argv[2], 
            'user': sys.argv[3], 
            'pass': sys.argv[4]
        })

        bludgeon2 = BdcReportGenerator(dbcred, sys.argv[5], sys.argv[6])
        bludgeon2.generateReport()

        # dbcred = dict({ 
        #     'host': '34.105.205.190', 
        #     'database': 'dashboard_service_v1', 
        #     'user': 'app', 
        #     'pass': 'password'
        # })

        # bludgeon2 = BdcReportGenerator(
        #     dbcred,
        #     'scripts/template/bdc_report_prototype.xlsx',
        #     'bdc-report-{0}.xlsx'.format(datetime.today().strftime('%d-%m-%Y'))
        # )
        # bludgeon2.generateReport()
    except (Exception) as e:
        print(e)
        # print(traceback.format_exc())


if __name__ == '__main__':
    generate_report()